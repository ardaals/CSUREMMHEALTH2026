from parser import parse_graphml, graphml_to_string
from pathlib import Path
from util import get_files_in_directory, output_sheet, copy_files_in_directory
from connectome import edge_attributes, node_attributes, edge_validation
from os import path
from plot import plot_connectome_3d
from openpyxl import Workbook
from connectome import connectivity_matrix
import pandas as pd
from wc_criticality import run_criticality_analysis
from network_control import run_network_control_analysis
from plot import plot_all_lever_connectivity_spearman


def create_connectivity_matrix(source_directory, output_directory, spectral_radius=False):
    files = get_files_in_directory(source_directory)
    for file in files:
        graph_str = graphml_to_string(path.join(source_directory,file))
        graph = parse_graphml(graph_str)
        W, nodes = connectivity_matrix(graph, spectral_radius=spectral_radius)
        df = pd.DataFrame(W, index=nodes, columns=nodes)
        # create output directory if it doesn't exist
        Path(output_directory).mkdir(parents=True, exist_ok=True)
        df.to_csv(path.join(output_directory, file + ".csv"))



def node_info(source_directory, output_directory):
    # parameters: source_directory is the directory where the graphml files are located
    # output_directory is the directory where the output spreadsheet will be saved
    # output_file_name is the name of the output spreadsheet, without the .xlsx extension
    wb = Workbook()
    ws1 = wb.active
    ws1.append(["Node Table", "Node ID", "Node Degree", "Position", "Region", "FreeSurfer Name", "Hemisphere", "Connected Nodes"])
    ws2 = wb.create_sheet("Edge Table")
    ws2.append(["Source Node ID", "Target Node ID", "Fiber Length Mean", "Fractional Anisotropy Mean", "Number of Fibers"])
    files = get_files_in_directory(source_directory)
    for file in files:
        graph_str = graphml_to_string(path.join(source_directory,file))
        graph = parse_graphml(graph_str)
        rows_node = node_attributes(graph)
        edge_info = edge_attributes(graph)
        for output in rows_node:
            ws1.append(output)
        for output in edge_info:
            ws2.append(output)
        print(f"{file}")
        output_sheet(output_directory, f"{file}_node_info", wb)

def edge_validate(source_directory, output_directory):
    wb = Workbook()
    ws1 = wb.active 
    ws1.append(["File Name", "Node Name"])
    files = get_files_in_directory(source_directory)
    Path(output_directory).mkdir(parents=True, exist_ok=True)
    for file in files:
        graph_str = graphml_to_string(path.join(source_directory,file))
        graph = parse_graphml(graph_str)
        rows_node = edge_validation(graph)
        ws1.append([file, rows_node])
        print(f"{file}")
    output_sheet(output_directory, "test_edge_validation", wb)


def clean_files(source_directory, output_directory):
    files = get_files_in_directory(source_directory)
    Path(output_directory).mkdir(parents=True, exist_ok=True)
    clean_files = []
    for file in files:
        graph_str = graphml_to_string(path.join(source_directory,file))
        graph = parse_graphml(graph_str)
        skip_file = False
        for node, data in graph.nodes(data=True):
            if graph.degree(node) == 0:
                skip_file = True
        if skip_file == True:
            continue
        else: 
            clean_files.append(path.join(source_directory,file))
    copy_files_in_directory(clean_files, output_directory)
    
        
        


def plot_graph(source_directory):
    files = get_files_in_directory(source_directory)
    for file in files:
        graph_str = graphml_to_string(path.join(source_directory,file))
        graph = parse_graphml(graph_str)
        plot_connectome_3d(graph)


def criticality_analysis(source_directory, output_directory, spectral_radius=False, set_noise=False):
    files = get_files_in_directory(source_directory)

    Path(output_directory).mkdir(parents=True, exist_ok=True)

    summary_rows = []

    for file in files:
        print(f"{file}")
        graph_str = graphml_to_string(path.join(source_directory, file))
        graph = parse_graphml(graph_str)

        W, nodes = connectivity_matrix(graph, spectral_radius=spectral_radius)

        output_prefix = Path(file).stem

        print("criticality")

        result = run_criticality_analysis(
            W=W,
            output_directory=output_directory,
            output_prefix=output_prefix,
            c5_min=0.0,
            c5_max=1000.0,
            num_c5=2001,
            spectral_radius = spectral_radius,
            set_noise = set_noise
        )

        summary_rows.append({
            "file": file,
            "c5_star": result["c5_star"],
            "jump_size": result["jump_size"],
            "summary_csv_path": result["summary_csv_path"],
            "regional_csv_path": result["regional_csv_path"],
            "plot_path": result["plot_path"],
        })

    summary_df = pd.DataFrame(summary_rows)
    summary_df.to_csv(
        path.join(output_directory, "criticality_all_connectomes_summary.csv"),
        index=False
    )


def criticality_and_network_control_analysis(
    source_directory,
    output_directory,
    spectral_radius=False,
    set_noise = False,
    T_control=1.0,
):
    files = get_files_in_directory(source_directory)

    Path(output_directory).mkdir(parents=True, exist_ok=True)

    summary_rows = []

    for file in files:
        print(f"{file}")

        graph_str = graphml_to_string(path.join(source_directory, file))
        graph = parse_graphml(graph_str)
        node_metadata = node_attributes(graph)
        # This is the only place W normalization happens.
        # network_control.py will NOT renormalize W or A.
        W, nodes = connectivity_matrix(
            graph,
            spectral_radius=spectral_radius,
        )

        output_prefix = Path(file).stem

        criticality_result = run_criticality_analysis(
            W=W,
            output_directory=output_directory,
            output_prefix=output_prefix,
            c5_min=0.0,
            c5_max=1000.0,
            num_c5=201,
            spectral_radius=spectral_radius,
            set_noise = set_noise
        )
        print("Done")

        network_control_df = run_network_control_analysis(
            W=W,
            nodes=nodes,
            criticality_result=criticality_result,
            output_directory=output_directory,
            output_prefix=output_prefix,
            spectral_radius=spectral_radius,
            T_control=T_control,
            fixed_point_T=250,
            fixed_point_dt=0.5,
            control="E",
            expm_version="eig",
            nctpy_dt=0.001,
            scale_energy_by_dt=True,
            save_arrays=False,
            node_metadata=node_metadata,
            save_excel=True,
            save_jacobian_excel_file=True,
        )


        network_control_csv_path = (
            Path(output_directory) / f"{output_prefix}_network_control_rankings.csv"
        )

        lever_connectivity_corr_df = plot_all_lever_connectivity_spearman(
            results_df=network_control_df,
            output_directory=output_directory,
            output_prefix=output_prefix,
            connectivity_metrics=[
                "weighted_degree",
                "binary_degree",
                "mean_connection_weight",
                "max_connection_weight",
            ],
            lever_rank_column="rank",
            use_connectivity_rank=True,
            annotate_top_n=5,
        )

        top_region = network_control_df.iloc[0]

        summary_rows.append({
            "file": file,
            "spectral_radius": bool(spectral_radius),
            "c5_star": criticality_result["c5_star"],
            "c5_pre": criticality_result["c5_pre"],
            "jump_size": criticality_result["jump_size"],
            "summary_csv_path": criticality_result["summary_csv_path"],
            "regional_csv_path": criticality_result["regional_csv_path"],
            "criticality_plot_path": criticality_result["plot_path"],
            "network_control_csv_path": str(network_control_csv_path),
            "top_lever_node": top_region["node"],
            "top_lever_energy": top_region["energy"],
            "top_lever_rank": top_region["rank"],
            "lambda_max_real": top_region["lambda_max_real"],
            "leading_eigenvalue_real": top_region["leading_eigenvalue_real"],
            "leading_eigenvalue_imag": top_region["leading_eigenvalue_imag"],
            "fixed_point_residual_norm": top_region["fixed_point_residual_norm"],
            "fixed_point_post_std": top_region["fixed_point_post_std"],
            "inversion_error": top_region["inversion_error"],
            "reconstruction_error": top_region["reconstruction_error"],
            "T_control": top_region["T_control"],
            "estimated_control_steps": top_region["estimated_control_steps"],
            "binary_degree": top_region["binary_degree"],
            "weighted_degree": top_region["weighted_degree"],
            "mean_connection_weight": top_region["mean_connection_weight"],
            "max_connection_weight": top_region["max_connection_weight"],
            "weighted_degree_rank": top_region["weighted_degree_rank"],
            "binary_degree_rank": top_region["binary_degree_rank"],       
        })

    summary_df = pd.DataFrame(summary_rows)

    summary_df.to_csv(
        path.join(output_directory, "criticality_and_network_control_summary.csv"),
        index=False,
    )

    
    return summary_df


if __name__ == '__main__':
    #put your code here

    pass
    # plot_graph("./data/dummy2")
    # node_info("./data/83node_413brains", "./spreadsheets")
    # clean_files("./data/83node_413brains", "./data/83node_413brains_onlyconnected")
    # edge_validate("./data/83node_413brains", "./spreadsheets/edge_validation")
    # create_connectivity_matrix("./data/dummy", "./data/dummy/connectivity_matrices_spectral", spectral_radius=True)
    criticality_analysis("./data/dummy3", "./data/dummy3/criticality3",spectral_radius=True, set_noise=False)
    # criticality_and_network_control_analysis(source_directory="./data/dummy3", output_directory="./data/dummy3/criticality_and_network_control_spectral", spectral_radius=True, set_noise=False)
